# Imports
from flask import Flask, render_template, request, redirect, url_for, send_from_directory
import os
import qrcode
from openpyxl import Workbook, load_workbook
from datetime import datetime
from PIL import Image
from pyzbar.pyzbar import decode

app = Flask(__name__)

# Function to generate a unique serial number using timestamp
def generate_serial():
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    return timestamp

# Function to write checklist data to Excel file
def write_to_excel(data):
    excel_file = 'Vehicle checklist data.xlsx'
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Serial Number', 'Vehicle Type', 'Vehicle Model', 'Vehicle Color', 'Repair Needed', 'Remarks', 'Update Time'])
    
    # Add a new column for update time
    data['update_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    ws.append([data['serial_number'], data['vehicleType'], data['vehicleModel'], data['vehicleColor'], data['repairNeeded'], data['remarks'], data['update_time']])
    wb.save(excel_file)

# Function to update QR code
def update_qr_code(data):
    qr_data = "\n".join([f"{key}: {value}" for key, value in data.items()])
    qr = qrcode.make(qr_data)
    qr.save(f'static/{data["serial_number"]}.png')

# Route to handle form submission and QR code generation
@app.route('/generate_qr', methods=['POST'])
def generate_qr():
    data = {
        'serial_number': generate_serial(),
        'vehicleType': request.form['vehicleType'],
        'vehicleModel': request.form['vehicleModel'],
        'vehicleColor': request.form['vehicleColor'],
        'repairNeeded': request.form['repairNeeded'],
        'remarks': request.form['remarks']
    }
    
    # Write data to Excel file
    write_to_excel(data)
    
    # Update QR code
    update_qr_code(data)
    
    return redirect(url_for('qr_display'))

# Route to handle QR code upload
@app.route('/upload_qr', methods=['POST'])
def upload_qr():
    if 'file' not in request.files:
        return render_template('error.html', message='No file part')
    file = request.files['file']
    if file.filename == '':
        return render_template('error.html', message='No selected file')
    if file:
        filename = file.filename
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        # Process the uploaded QR code image
        data = process_qr_code_image(filename)
        if data:
            # Update Excel file and QR code
            write_to_excel(data)
            update_qr_code(data)
            return redirect(url_for('qr_display'))
        else:
            # Handle invalid QR code
            return render_template('error.html', message='Invalid QR code')
    else:
        # Handle no file uploaded
        return render_template('error.html', message='No file uploaded')

# Function to process uploaded QR code image
def process_qr_code_image(filename):
    try:
        # Open the image file
        with open(filename, 'rb') as image_file:
            # Decode the QR code
            decoded_objects = decode(Image.open(image_file))
            if decoded_objects:
                # Extract data from the decoded QR code
                decoded_data = decoded_objects[0].data.decode('utf-8')
                # Split the data into key-value pairs
                data = dict(item.split(': ') for item in decoded_data.split('\n'))
                return data
            else:
                return None
    except Exception as e:
        print(e)
        return None

# Route to display QR code
@app.route('/qr_display')
def qr_display():
    return render_template('qr_display.html')

# Route to serve uploaded files
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

if __name__ == '__main__':
    app.run(debug=True)
