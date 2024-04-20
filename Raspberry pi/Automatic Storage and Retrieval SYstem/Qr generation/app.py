from flask import Flask, render_template, request, redirect, url_for
import os
import qrcode
from openpyxl import Workbook, load_workbook
from datetime import datetime

app = Flask(__name__)

# Function to generate a unique serial number using timestamp
def generate_serial():
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    return f'{timestamp}'

# Function to write checklist data to Excel file
def write_to_excel(data):
    excel_file = 'Vehicle checklist data.xlsx'
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Serial Number', 'Vehicle Type', 'Vehicle Model', 'Vehicle Color', 'Repair Needed', 'Remarks'])
    
    ws.append([data['serial_number'], data['vehicleType'], data['vehicleModel'], data['vehicleColor'], data['repairNeeded'], data['remarks']])
    wb.save(excel_file)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate_qr', methods=['POST'])
def generate_qr():
    data = {
        'serial_number': generate_serial(),
        'vehicleType': request.form['vehicleType'],
        'vehicleModel': request.form['vehicleModel'],
        'vehicleColor': request.form['vehicleColor'],
        'repairNeeded': request.form['repairNeeded'],
        'remarks': request.form['remarks']
    }
    
    # Write data to Excel file
    write_to_excel(data)
    
    qr_data = "\n".join([f"{key}: {value}" for key, value in data.items()])
    
    # Create the static directory if it doesn't exist
    if not os.path.exists('static'):
        os.makedirs('static')
    
    qr = qrcode.make(qr_data)
    qr.save(f'static/{data["serial_number"]}.png')
    
    return redirect(url_for('qr_display'))

@app.route('/qr_display')
def qr_display():
    return render_template('qr_display.html')

if __name__ == '__main__':
    app.run(debug=True)
