from pyzbar.pyzbar import decode
from PIL import Image
from openpyxl import load_workbook
import os

# Function to decode QR code from image
def decode_qr_code(image_filename):
    try:
        with open(image_filename, 'rb') as image_file:
            image = Image.open(image_file)
            decoded_objects = decode(image)
            if decoded_objects:
                return decoded_objects[0].data.decode('utf-8')
            else:
                return None
    except Exception as e:
        print("Error decoding QR code:", e)
        return None

# Function to read data from Excel file
def read_from_excel(excel_filename):
    try:
        wb = load_workbook(filename=excel_filename)
        ws = wb.active
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        return data
    except Exception as e:
        print("Error reading from Excel file:", e)
        return None

# Example usage
excel_directory = "Automatic_Storage_and_Retrieval_System"
excel_filename = "vehicle_Database.xlsx"
excel_full_path = os.path.join(excel_directory, excel_filename)

# Read data from Excel file
excel_data = read_from_excel(excel_filename)

if excel_data:
    # Extract QR code filename from cell C3
    qr_code_filename = excel_data[2][0]  # Assuming QR code filename is in cell C3

    # Decode QR code from the QR code image filename stored in Excel
    decoded_data = decode_qr_code(qr_code_filename)

    if decoded_data:
        print("Decoded QR code data:", decoded_data)
    else:
        print("No QR code found or could not be decoded.")
else:
    print("Error reading data from Excel file.")


 